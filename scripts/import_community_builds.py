from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = BASE_DIR / "data" / "community_builds.json"

MAIN_SHEET = "阵容主表"
CARDS_SHEET = "阵容卡牌"
OPTIONS_SHEET = "选项"
CARD_REF_SHEET = "卡牌参考"
SOURCES_SHEET = "阵容来源"

STAGE_COLUMNS = [
    ("适用前期", "early"),
    ("适用中期", "mid"),
    ("适用后期", "late"),
]

CARD_ROLE_FIELDS = {
    "core_cards",
    "transition_cards",
    "optional_cards",
}


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)
        file.write("\n")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_list(value: Any, separator: str) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return [item.strip() for item in text.split(separator) if item.strip()]


def yes(value: Any) -> bool:
    return clean_text(value) == "是"


def int_or_none(value: Any) -> int | None:
    if value in (None, ""):
        return None
    return int(value)


def float_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def headers(ws) -> dict[str, int]:
    return {
        clean_text(cell.value): cell.column
        for cell in ws[1]
        if clean_text(cell.value)
    }


def row_value(ws, row: int, header_map: dict[str, int], name: str) -> Any:
    col = header_map.get(name)
    if col is None:
        return None
    return ws.cell(row, col).value


def require_sheets(wb) -> None:
    required = [MAIN_SHEET, CARDS_SHEET, OPTIONS_SHEET, CARD_REF_SHEET]
    missing = [name for name in required if name not in wb.sheetnames]
    if missing:
        raise ValueError(f"模板缺少工作表：{', '.join(missing)}")


def option_maps(wb) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    options = wb[OPTIONS_SHEET]
    card_ref = wb[CARD_REF_SHEET]

    hero_map = {
        clean_text(options.cell(row, 1).value): clean_text(options.cell(row, 2).value)
        for row in range(2, options.max_row + 1)
        if clean_text(options.cell(row, 1).value)
    }
    role_map = {
        clean_text(options.cell(row, 5).value): clean_text(options.cell(row, 6).value)
        for row in range(2, options.max_row + 1)
        if clean_text(options.cell(row, 5).value)
    }
    card_map = {
        clean_text(card_ref.cell(row, 1).value): clean_text(card_ref.cell(row, 2).value)
        for row in range(2, card_ref.max_row + 1)
        if clean_text(card_ref.cell(row, 1).value)
    }
    return hero_map, role_map, card_map


def parse_main_builds(wb, hero_map: dict[str, str]) -> tuple[dict[str, dict[str, Any]], list[str]]:
    ws = wb[MAIN_SHEET]
    header_map = headers(ws)
    builds: dict[str, dict[str, Any]] = {}
    errors: list[str] = []

    for row in range(2, ws.max_row + 1):
        build_id = clean_text(row_value(ws, row, header_map, "build_id*"))
        if not build_id:
            continue

        if build_id in builds:
            errors.append(f"{MAIN_SHEET}!A{row}：build_id 重复：{build_id}")
            continue

        hero_cn = clean_text(row_value(ws, row, header_map, "英雄中文*"))
        hero = hero_map.get(hero_cn, "")
        display_name = clean_text(row_value(ws, row, header_map, "display_name*"))
        day_start = int_or_none(row_value(ws, row, header_map, "day_start*"))
        day_end = int_or_none(row_value(ws, row, header_map, "day_end(空=不限)"))

        if not hero:
            errors.append(f"{MAIN_SHEET}!B{row}：英雄无法映射：{hero_cn}")
        if not display_name:
            errors.append(f"{MAIN_SHEET}!D{row}：display_name 不能为空")
        if day_start is None:
            errors.append(f"{MAIN_SHEET}!H{row}：day_start 不能为空")

        applicable_stages = [
            stage
            for column_name, stage in STAGE_COLUMNS
            if yes(row_value(ws, row, header_map, column_name))
        ]
        if not applicable_stages:
            errors.append(f"{MAIN_SHEET}!E:G{row}：至少选择一个适用阶段")

        build: dict[str, Any] = {
            "hero": hero,
            "display_name": display_name,
            "applicable_stages": applicable_stages,
            "day_range": [day_start, day_end],
            "build_summary": clean_text(row_value(ws, row, header_map, "build_summary")),
            "match_notes": split_list(row_value(ws, row, header_map, "match_notes(|分隔)"), "|"),
            "pilot_tips": split_list(row_value(ws, row, header_map, "经验与Tips(|分隔)"), "|"),
            "core_cards": [],
            "transition_cards": [],
            "optional_cards": [],
            "wanted_tags": split_list(row_value(ws, row, header_map, "wanted_tags(,分隔)"), ","),
            "event_priorities": split_list(row_value(ws, row, header_map, "event_priorities(,分隔)"), ","),
            "avoid_events": split_list(row_value(ws, row, header_map, "avoid_events(,分隔)"), ","),
        }

        source_type = clean_text(row_value(ws, row, header_map, "source_type"))
        confidence = clean_text(row_value(ws, row, header_map, "confidence"))
        if source_type:
            build["source_type"] = source_type
        if confidence:
            build["confidence"] = confidence

        builds[build_id] = build

    return builds, errors


def parse_build_cards(
    wb,
    builds: dict[str, dict[str, Any]],
    role_map: dict[str, str],
    card_map: dict[str, str],
) -> tuple[dict[str, list[dict[str, Any]]], list[str]]:
    ws = wb[CARDS_SHEET]
    header_map = headers(ws)
    errors: list[str] = []
    frequencies: dict[str, list[dict[str, Any]]] = defaultdict(list)
    seen_cards: dict[str, set[str]] = defaultdict(set)

    for row in range(2, ws.max_row + 1):
        build_id = clean_text(row_value(ws, row, header_map, "build_id*"))
        card_cn = clean_text(row_value(ws, row, header_map, "卡牌中文名*"))
        role_cn = clean_text(row_value(ws, row, header_map, "定位中文*"))
        if not any([build_id, card_cn, role_cn]):
            continue

        if build_id not in builds:
            errors.append(f"{CARDS_SHEET}!A{row}：build_id 不存在：{build_id}")
            continue

        card_name = card_map.get(card_cn, "")
        role_field = role_map.get(role_cn, "")
        if not card_name:
            errors.append(f"{CARDS_SHEET}!B{row}：卡牌无法映射：{card_cn}")
            continue
        if role_field not in CARD_ROLE_FIELDS:
            errors.append(f"{CARDS_SHEET}!D{row}：定位无法映射：{role_cn}")
            continue

        if card_name not in seen_cards[build_id]:
            builds[build_id][role_field].append(card_name)
            seen_cards[build_id].add(card_name)

        count = int_or_none(row_value(ws, row, header_map, "count(可选)"))
        frequency = float_or_none(row_value(ws, row, header_map, "frequency(可选)"))
        if count is not None or frequency is not None:
            item: dict[str, Any] = {"name": card_name}
            if count is not None:
                item["count"] = count
            if frequency is not None:
                item["frequency"] = frequency
            frequencies[build_id].append(item)

    return frequencies, errors


def parse_sources(wb, builds: dict[str, dict[str, Any]]) -> list[str]:
    if SOURCES_SHEET not in wb.sheetnames:
        return []

    ws = wb[SOURCES_SHEET]
    header_map = headers(ws)
    errors: list[str] = []
    sources: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in range(2, ws.max_row + 1):
        build_id = clean_text(row_value(ws, row, header_map, "build_id*"))
        url = clean_text(row_value(ws, row, header_map, "url*"))
        if not any([build_id, url]):
            continue

        if build_id not in builds:
            errors.append(f"{SOURCES_SHEET}!A{row}：build_id 不存在：{build_id}")
            continue

        source = {
            "id": clean_text(row_value(ws, row, header_map, "run_id")),
            "url": url,
            "player": clean_text(row_value(ws, row, header_map, "player")),
            "hero": clean_text(row_value(ws, row, header_map, "hero")),
            "record": clean_text(row_value(ws, row, header_map, "record")),
            "result": clean_text(row_value(ws, row, header_map, "result")),
            "title": clean_text(row_value(ws, row, header_map, "title")),
            "imported_at_utc": row_value(ws, row, header_map, "imported_at_utc"),
        }
        sources[build_id].append({key: value for key, value in source.items() if value not in ("", None)})

    for build_id, build_sources in sources.items():
        builds[build_id]["source_runs"] = build_sources

    return errors


def import_workbook(path: Path) -> tuple[dict[str, Any], list[str]]:
    wb = load_workbook(path, data_only=False, read_only=False)
    require_sheets(wb)

    hero_map, role_map, card_map = option_maps(wb)
    builds, errors = parse_main_builds(wb, hero_map)
    frequencies, card_errors = parse_build_cards(wb, builds, role_map, card_map)
    errors.extend(card_errors)
    errors.extend(parse_sources(wb, builds))

    for build_id, frequency_items in frequencies.items():
        builds[build_id]["card_frequencies"] = frequency_items

    for build_id, build in builds.items():
        if not any(build[field] for field in CARD_ROLE_FIELDS):
            errors.append(f"{build_id}：至少需要填写一张核心/过渡/可选卡")

    return builds, errors


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="把社区阵容 Excel 模板转换为 community_builds.json。")
    parser.add_argument("input", type=Path, help="社区阵容录入模板 .xlsx 路径")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="独立输出 JSON 路径")
    parser.add_argument("--allow-errors", action="store_true", help="有校验错误时仍然输出")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    builds, errors = import_workbook(args.input)

    print("社区阵容转换结果：")
    print(f"- 输入：{args.input}")
    print(f"- 读取阵容：{len(builds)}")
    print(f"- 校验问题：{len(errors)}")
    for error in errors[:50]:
        print(f"  - {error}")
    if len(errors) > 50:
        print(f"  ... 还有 {len(errors) - 50} 个问题")

    if errors and not args.allow_errors:
        raise SystemExit("存在校验问题，已停止输出。修正表格后重试，或加 --allow-errors 强制输出。")

    write_json(args.output, builds)
    print(f"- 已输出独立 JSON：{args.output}")



if __name__ == "__main__":
    main()
