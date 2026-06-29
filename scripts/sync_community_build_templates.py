from __future__ import annotations

import json
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
TEMPLATE_PATH = (
    BASE_DIR
    / "outputs"
    / "community_build_template"
    / "社区阵容录入模板_v4_完整卡牌_可用版.xlsx"
)
REFERENCE_HEADERS = ["中文卡名", "英文内部名", "英雄", "类型", "标签"]


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig") as file:
        data = json.load(file)
    return data if isinstance(data, dict) else {}


def load_candidates() -> list[list[Any]]:
    translations = load_json(DATA_DIR / "translations_zh_cn.json")
    by_name = translations.get("by_name", {})
    by_id = translations.get("by_id", {})
    candidates: list[list[Any]] = []

    for filename in ("cards_generated.json", "skills_generated.json"):
        for key, card in load_json(DATA_DIR / filename).items():
            if not isinstance(card, dict) or card.get("type") not in {"Item", "Skill"}:
                continue

            english_name = str(card.get("name") or card.get("internal_name") or key)
            card_id = str(card.get("id") or "")
            chinese_name = (
                by_id.get(card_id)
                or by_name.get(english_name)
                or by_name.get(key)
                or english_name
            )
            heroes = card.get("heroes") or []
            hero_text = (
                " / ".join(str(hero) for hero in heroes if hero)
                if isinstance(heroes, list)
                else str(card.get("hero") or "")
            )
            tags = card.get("tags") or []
            tag_text = ", ".join(str(tag) for tag in tags) if isinstance(tags, list) else ""
            candidates.append(
                [
                    chinese_name,
                    card.get("internal_name") or card.get("source_id") or english_name,
                    hero_text,
                    card.get("type"),
                    tag_text,
                ]
            )

    return sorted(
        candidates,
        key=lambda row: (0 if row[3] == "Item" else 1, str(row[0])),
    )


def find_sheet(workbook, headers: list[str]):
    for worksheet in workbook.worksheets:
        values = [worksheet.cell(1, column).value for column in range(1, len(headers) + 1)]
        if values == headers:
            return worksheet
    raise ValueError(f"找不到表头：{headers}")


def copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    for column in range(1, 6):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        target._style = copy(source._style)
        target.number_format = source.number_format


def sync_template(path: Path = TEMPLATE_PATH) -> tuple[int, int]:
    candidates = load_candidates()
    workbook = load_workbook(path)
    reference_sheet = find_sheet(workbook, REFERENCE_HEADERS)
    cards_sheet = workbook.worksheets[2]

    old_last_row = reference_sheet.max_row
    for row in range(2, old_last_row + 1):
        for column in range(1, 6):
            reference_sheet.cell(row, column).value = None

    for row, values in enumerate(candidates, start=2):
        if row > old_last_row:
            copy_row_style(reference_sheet, old_last_row, row)
        for column, value in enumerate(values, start=1):
            reference_sheet.cell(row, column).value = value

    last_row = len(candidates) + 1
    reference_title = reference_sheet.title.replace("'", "''")
    workbook.defined_names.pop("ref_cards", None)
    workbook.defined_names.add(
        DefinedName(
            "ref_cards",
            attr_text=f"'{reference_title}'!$A$2:$A${last_row}",
        )
    )

    for row in range(2, cards_sheet.max_row + 1):
        cards_sheet.cell(row, 3).value = (
            f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},'
            f"'{reference_title}'!$A$2:$B${last_row},2,FALSE),\"\"))"
        )

    workbook.save(path)
    skill_count = sum(1 for row in candidates if row[3] == "Skill")
    return len(candidates), skill_count


def main() -> None:
    candidate_count, skill_count = sync_template()
    print(f"updated: {TEMPLATE_PATH.name}")
    print(f"candidates: {candidate_count}, skills: {skill_count}")


if __name__ == "__main__":
    main()
